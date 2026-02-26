import pandas as pd
import re
import os
import json
import concurrent.futures
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, create_model
from google import genai


def detect_excel_header_row(uploaded_file) -> Optional[int]:
    """
    Detect the header row in an Excel file using formatting and content analysis.
    
    Args:
        uploaded_file: File object from Streamlit's file_uploader
        
    Returns:
        Row number (0-based) where headers are found, or None if not detected
    """
    if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        return None
    
    try:
        # Reset file pointer
        uploaded_file.seek(0)
        
        # Load workbook with openpyxl to analyze formatting
        workbook = load_workbook(uploaded_file, data_only=False)
        worksheet = workbook.active
        
        # Get merged cell ranges
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        # Analyze first 10 rows for header detection
        max_rows_to_check = min(10, worksheet.max_row)
        
        for row_num in range(1, max_rows_to_check + 1):
            # Skip if this row is part of a merged cell
            if _is_row_in_merged_cell(row_num, merged_ranges):
                continue
            
            # Get row data
            row_data = []
            bold_count = 0
            total_cells = 0
            
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell_value = cell.value
                
                # Skip empty cells
                if cell_value is None or str(cell_value).strip() == '':
                    continue
                
                total_cells += 1
                row_data.append(str(cell_value).strip())
                
                # Check if cell is bold
                if cell.font and cell.font.bold:
                    bold_count += 1
            
            # Skip rows with too few cells
            if total_cells < 2:
                continue
            
            # Calculate bold percentage
            bold_percentage = bold_count / total_cells if total_cells > 0 else 0
            
            # Check if this looks like a header row
            if _is_likely_header_row(row_data, bold_percentage):
                return row_num - 1  # Convert to 0-based index for pandas
        
        return None
        
    except Exception as e:
        # If openpyxl analysis fails, fall back to content-based detection
        print(f"Excel formatting analysis failed: {e}")
        return None


def _is_row_in_merged_cell(row_num: int, merged_ranges) -> bool:
    """Check if a row is part of any merged cell range."""
    for merged_range in merged_ranges:
        if merged_range.min_row <= row_num <= merged_range.max_row:
            return True
    return False


def _is_likely_header_row(row_data: List[str], bold_percentage: float) -> bool:
    """
    Determine if a row is likely to contain headers based on content and formatting.
    
    Args:
        row_data: List of cell values in the row
        bold_percentage: Percentage of cells that are bold
        
    Returns:
        True if this row is likely to contain headers
    """
    if len(row_data) < 2:
        return False
    
    # Check content-based indicators
    content_score = 0
    
    # Most cells should be strings (not numbers)
    string_count = sum(1 for val in row_data if isinstance(val, str) and not val.isdigit())
    if string_count / len(row_data) >= 0.7:
        content_score += 1
    
    # Check for common header patterns
    header_indicators = ['name', 'id', 'email', 'date', 'time', 'question', 'response', 'answer', 'category', 'type', 'status']
    header_matches = sum(1 for val in row_data if any(indicator in val.lower() for indicator in header_indicators))
    if header_matches > 0:
        content_score += 1
    
    # Check for reasonable header length (not too short, not too long)
    avg_length = sum(len(val) for val in row_data) / len(row_data)
    if 3 <= avg_length <= 50:
        content_score += 1
    
    # Formatting-based indicators
    formatting_score = 0
    if bold_percentage >= 0.5:  # At least 50% of cells are bold
        formatting_score += 2
    elif bold_percentage >= 0.3:  # At least 30% of cells are bold
        formatting_score += 1
    
    # Combined score: need at least 2 points total
    total_score = content_score + formatting_score
    return total_score >= 2


def load_flexible_data(uploaded_file) -> pd.DataFrame:
    """
    Load data from uploaded file with flexible header detection.
    
    Args:
        uploaded_file: File object from Streamlit's file_uploader
        
    Returns:
        DataFrame with properly detected headers
    """
    # Load the file based on extension
    if uploaded_file.name.lower().endswith('.csv'):
        # Try different skiprow values to find headers
        for skip_rows in [0, 1, 2, 3, 4]:
            try:
                df = pd.read_csv(uploaded_file, skiprows=skip_rows)
                if _is_valid_header_row(df.columns):
                    return df
            except:
                continue
        # If all fail, use default
        uploaded_file.seek(0)  # Reset file pointer
        return pd.read_csv(uploaded_file, skiprows=3)
        
    elif uploaded_file.name.lower().endswith('.json'):
        return pd.read_json(uploaded_file)
    else:
        # Excel files - use enhanced detection
        detected_header_row = detect_excel_header_row(uploaded_file)
        
        if detected_header_row is not None:
            # Use the detected header row
            uploaded_file.seek(0)  # Reset file pointer
            try:
                df = pd.read_excel(uploaded_file, sheet_name=0, skiprows=detected_header_row)
                if _is_valid_header_row(df.columns):
                    print(f"Successfully detected headers in row {detected_header_row + 1} using Excel formatting analysis")
                    return df
            except Exception as e:
                print(f"Failed to load with detected header row {detected_header_row}: {e}")
        
        # Fall back to original method if enhanced detection fails
        for skip_rows in [0, 1, 2, 3, 4]:
            try:
                df = pd.read_excel(uploaded_file, sheet_name=0, skiprows=skip_rows)
                if _is_valid_header_row(df.columns):
                    return df
            except:
                continue
        
        # If all fail, use default
        uploaded_file.seek(0)  # Reset file pointer
        return pd.read_excel(uploaded_file, sheet_name=0, skiprows=3)


def _is_valid_header_row(columns) -> bool:
    """
    Check if a row contains valid column headers.
    
    Args:
        columns: List of column values to check
        
    Returns:
        True if this looks like a valid header row
    """
    if len(columns) < 2:  # Need at least 2 columns
        return False
    
    # Check if most columns are strings (not NaN or numbers)
    string_count = sum(1 for col in columns if isinstance(col, str) and str(col).strip())
    return string_count >= len(columns) * 0.7  # At least 70% should be strings


def get_text_columns(df: pd.DataFrame) -> List[str]:
    """
    Get list of columns that are suitable for text categorisation.
    
    Args:
        df: DataFrame to analyze
        
    Returns:
        List of column names that contain text data suitable for categorisation
    """
    text_columns = []
    
    for col in df.columns:
        # Check if column is object/string type
        if df[col].dtype == 'object':
            # Skip columns that are already boolean (like _customer columns)
            if col.endswith('_customer') or col.endswith('_child'):
                continue
            
            # Skip columns that look like they contain structured data
            if col.startswith('Q(') and '[' in col:
                continue
                
            # Check if column contains mostly text (not just numbers)
            non_null_values = df[col].dropna()
            if len(non_null_values) > 0:
                # Sample a few values to check if they're text-like
                sample_values = non_null_values.head(10)
                text_like_count = 0
                for val in sample_values:
                    val_str = str(val).strip()
                    # Consider it text-like if it's longer than 10 chars or contains spaces
                    if len(val_str) > 10 or ' ' in val_str:
                        text_like_count += 1
                
                # If at least 50% of sample values look like text, include this column
                if text_like_count / len(sample_values) >= 0.5:
                    text_columns.append(col)
    
    return text_columns


def estimate_categorisation_cost(
    text_token_count: int,
    num_responses: int,
    categories: List[str],
    api_key: Optional[str] = None
) -> Dict[str, float]:
    """
    Estimate the cost for categorisation based on token counts.
    
    Args:
        text_token_count: Total tokens for all text responses in the column
        num_responses: Number of responses to categorise
        categories: List of category labels
        api_key: Optional API key. If not provided, will try to get from GEMINI_API_KEY environment variable.
    
    Returns:
        Dictionary with 'input_cost', 'output_cost', and 'total_cost' in USD
    """
    try:
        client = _get_genai_client(api_key)
        
        # Estimate system prompt tokens
        system_prompt = create_system_prompt(categories)
        system_prompt_response = client.models.count_tokens(
            model='gemini-2.0-flash-lite',
            contents=[{"role": "user", "parts": [{"text": system_prompt}]}]
        )
        system_prompt_tokens = system_prompt_response.total_tokens
        
        # Estimate schema tokens (JSON schema for Pydantic model)
        # Rough estimate: ~50 base tokens + ~10 per category
        schema_tokens = 50 + (len(categories) * 10)
        
        # Average tokens per response (text data tokens / number of responses)
        avg_text_tokens_per_response = text_token_count / num_responses if num_responses > 0 else 0
        
        # Input tokens per API call = text tokens + system prompt + schema
        input_tokens_per_call = avg_text_tokens_per_response + system_prompt_tokens + schema_tokens
        
        # Total input tokens = input per call × number of responses
        total_input_tokens = input_tokens_per_call * num_responses
        
        # Output tokens: 10 tokens per category per response
        total_output_tokens = num_responses * len(categories) * 10
        
        # Calculate costs
        # Input: $0.075 per 1M tokens
        # Output: $0.30 per 1M tokens
        input_cost = (total_input_tokens / 1_000_000) * 0.075
        output_cost = (total_output_tokens / 1_000_000) * 0.30
        total_cost = input_cost + output_cost
        
        return {
            'input_cost': input_cost,
            'output_cost': output_cost,
            'total_cost': total_cost,
            'input_tokens': total_input_tokens,
            'output_tokens': total_output_tokens
        }
        
    except Exception as e:
        print(f"Error estimating cost: {e}")
        return {
            'input_cost': 0.0,
            'output_cost': 0.0,
            'total_cost': 0.0,
            'input_tokens': 0,
            'output_tokens': 0
        }


def count_tokens_for_column(df: pd.DataFrame, column_name: str, api_key: Optional[str] = None) -> Optional[int]:
    """
    Count tokens for all text in a column using Gemini API token counter.
    
    Args:
        df: DataFrame containing the column
        column_name: Name of the column to count tokens for
        api_key: Optional API key. If not provided, will try to get from GEMINI_API_KEY environment variable.
    
    Returns:
        Total token count, or None if counting fails
    """
    if column_name not in df.columns:
        return None
    
    try:
        client = _get_genai_client(api_key)
        
        # Get all non-null text values
        text_values = df[column_name].dropna().astype(str).tolist()
        
        if not text_values:
            return 0
        
        # Combine all text (this is an approximation - actual API calls will have system prompt too)
        # But for estimation purposes, we'll count the text data
        combined_text = "\n".join(text_values)
        
        # Count tokens using Gemini API
        response = client.models.count_tokens(
            model='gemini-2.0-flash-lite',
            contents=[{"role": "user", "parts": [{"text": combined_text}]}]
        )
        
        return response.total_tokens
        
    except Exception as e:
        print(f"Error counting tokens: {e}")
        return None


def preview_column_data(df: pd.DataFrame, column_name: str, num_samples: int = 10) -> Dict[str, Any]:
    """
    Get preview data for a text column to help users understand the content.
    
    Args:
        df: DataFrame containing the column
        column_name: Name of the column to preview
        num_samples: Number of sample responses to return
        
    Returns:
        Dictionary with preview information
    """
    if column_name not in df.columns:
        return {"error": f"Column '{column_name}' not found"}
    
    # Get non-null values
    non_null_values = df[column_name].dropna()
    
    if len(non_null_values) == 0:
        return {
            "total_responses": 0,
            "non_null_responses": 0,
            "sample_responses": [],
            "average_length": 0
        }
    
    # Get sample responses
    sample_responses = non_null_values.head(num_samples).tolist()
    
    # Calculate average length
    lengths = [len(str(val)) for val in non_null_values]
    avg_length = sum(lengths) / len(lengths) if lengths else 0
    
    return {
        "total_responses": len(df),
        "non_null_responses": len(non_null_values),
        "sample_responses": sample_responses,
        "average_length": round(avg_length, 1)
    }


def _get_genai_client(api_key: Optional[str] = None):
    """
    Get initialized Google GenAI client using API key.
    
    Args:
        api_key: Optional API key. If not provided, will try to get from GEMINI_API_KEY environment variable.
    
    Returns:
        Initialized genai.Client instance
    
    Raises:
        ValueError: If no API key is available
    """
    if api_key is None:
        api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY not provided and environment variable not set")
    return genai.Client(api_key=api_key)


def create_classification_model(categories: List[str]) -> BaseModel:
    """
    Create a dynamic Pydantic model based on user-specified categories.
    
    Args:
        categories: List of category names (strings)
    
    Returns:
        A Pydantic model class with boolean fields for each category
    """
    # Create field definitions for each category
    field_definitions = {category: (bool, False) for category in categories}
    
    # Create the model dynamically
    model_class = create_model(
        'DynamicClassification',
        **field_definitions,
        __base__=BaseModel
    )
    
    return model_class


def create_system_prompt(categories: List[str]) -> str:
    """
    Create a dynamic system prompt based on the categories.
    """
    category_list = "\n".join([f"- {category}" for category in categories])
    
    return f"""
You will be given a text snippet and your response will classify it into the following categories:

{category_list}

For each category, respond with true if the text belongs to that category, false otherwise.

A text can belong to multiple categories or none at all.
"""


def classify_text(
    text_snippet: str, 
    categories: List[str], 
    system_prompt: Optional[str] = None, 
    ClassificationModel: Optional[BaseModel] = None, 
    model_name: str = 'gemini-2.0-flash-lite',
    client: Optional[genai.Client] = None
) -> Dict[str, bool]:
    """
    Classify a text snippet into categories using Google GenAI.
    
    Args:
        text_snippet: The text to classify
        categories: List of category names
        system_prompt: Optional pre-generated system prompt
        ClassificationModel: Optional pre-generated Pydantic model
        model_name: Name of the model to use
        client: Optional GenAI client (will create if not provided)
    
    Returns:
        Dictionary mapping category names to boolean values
    """
    if ClassificationModel is None:
        ClassificationModel = create_classification_model(categories)
    if system_prompt is None:
        system_prompt = create_system_prompt(categories)
    if client is None:
        client = _get_genai_client()
    
    response = client.models.generate_content(
        model=model_name,
        contents=[
            {"role": "user", "parts": [{"text": text_snippet}]}
        ],
        config={
            "response_mime_type": "application/json",
            "response_schema": ClassificationModel,
            "temperature": 0.0,
            "system_instruction": system_prompt,
        }
    )
    response_json = json.loads(response.text)
    return response_json


def _classify_row(
    row: pd.Series,
    text_column: str,
    categories: List[str],
    system_prompt: str,
    ClassificationModel: BaseModel,
    client: genai.Client
) -> Tuple[int, Optional[Dict[str, bool]]]:
    """
    Worker function to classify a single row.
    
    Returns a tuple of (index, classification_results).
    """
    text = row[text_column]
    try:
        classification = classify_text(
            text_snippet=text,
            categories=categories,
            system_prompt=system_prompt,
            ClassificationModel=ClassificationModel,
            client=client
        )
        return (row.name, classification)
    except Exception as e:
        print(f"Error classifying row {row.name}: {e}")
        return (row.name, None)


def categorise_responses(
    df: pd.DataFrame, 
    column_name: str, 
    categories: List[str],
    max_workers: int = 10,
    progress_callback=None,
    api_key: Optional[str] = None
) -> pd.DataFrame:
    """
    Categorise text responses using parallel processing with Google GenAI.
    
    Args:
        df: DataFrame containing the text column
        column_name: Name of the column to categorise
        categories: List of category labels
        max_workers: Number of parallel workers (default: 10)
        progress_callback: Optional callback function for progress updates (takes progress as float 0-1)
        api_key: Optional API key. If not provided, will try to get from GEMINI_API_KEY environment variable.
    
    Returns:
        DataFrame with new boolean columns for each category
    """
    result_df = df.copy()
    
    # Create safe column names for each category
    category_columns = {}
    for category in categories:
        safe_category = category.lower().replace(' ', '_').replace('-', '_')
        safe_category = re.sub(r'[^a-z0-9_]', '', safe_category)
        new_column_name = f"{column_name}_{safe_category}"
        category_columns[category] = new_column_name
        # Initialize all columns with False
        result_df[new_column_name] = False
    
    # Set up classifier components (create once, reuse for all rows)
    system_prompt = create_system_prompt(categories)
    ClassificationModel = create_classification_model(categories)
    client = _get_genai_client(api_key)
    
    # Filter out rows with NaN values in the text column
    valid_rows = result_df.dropna(subset=[column_name])
    
    if len(valid_rows) == 0:
        return result_df
    
    total_rows = len(valid_rows)
    print(f"Processing {total_rows} valid rows out of {len(result_df)} total rows using {max_workers} workers")
    
    # Use ThreadPoolExecutor for parallel processing
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_row = {
            executor.submit(
                _classify_row,
                row,
                column_name,
                categories,
                system_prompt,
                ClassificationModel,
                client
            ): idx for idx, (_, row) in enumerate(valid_rows.iterrows())
        }
        
        # Process results as they complete
        completed = 0
        for future in concurrent.futures.as_completed(future_to_row):
            completed += 1
            if progress_callback:
                progress_callback(completed / total_rows)
            
            index, classification = future.result()
            if classification:
                for category in categories:
                    result_df.at[index, category_columns[category]] = classification.get(category, False)
    
    return result_df


def get_categorisation_summary(df: pd.DataFrame, column_name: str, categories: List[str]) -> Dict[str, Any]:
    """
    Get summary statistics for categorised data.
    
    Args:
        df: DataFrame with categorised columns
        column_name: Original text column name
        categories: List of category labels
        
    Returns:
        Dictionary with summary statistics
    """
    summary = {
        "total_responses": len(df),
        "categories": {}
    }
    
    for category in categories:
        safe_category = category.lower().replace(' ', '_').replace('-', '_')
        safe_category = re.sub(r'[^a-z0-9_]', '', safe_category)
        new_column_name = f"{column_name}_{safe_category}"
        
        if new_column_name in df.columns:
            category_count = df[new_column_name].sum()
            category_percentage = (category_count / len(df)) * 100 if len(df) > 0 else 0
            
            summary["categories"][category] = {
                "count": int(category_count),
                "percentage": round(category_percentage, 1)
            }
    
    return summary


def get_categorisation_summary_with_breakdown(
    df: pd.DataFrame, 
    column_name: str, 
    categories: List[str],
    breakdown_column: str
) -> Dict[str, Any]:
    """
    Get summary statistics for categorised data with breakdown by another column.
    
    Args:
        df: DataFrame with categorised columns
        column_name: Original text column name
        categories: List of category labels
        breakdown_column: Column name to break down by
        
    Returns:
        Dictionary with summary statistics broken down by breakdown_column values
    """
    summary = {
        "total_responses": len(df),
        "categories": {},
        "breakdown_values": []
    }
    
    # Get unique values in breakdown column
    breakdown_values = sorted(df[breakdown_column].dropna().unique().tolist())
    summary["breakdown_values"] = [str(v) for v in breakdown_values]
    
    # Calculate totals for each breakdown value
    breakdown_totals = {}
    for value in breakdown_values:
        breakdown_totals[str(value)] = len(df[df[breakdown_column] == value])
    
    for category in categories:
        safe_category = category.lower().replace(' ', '_').replace('-', '_')
        safe_category = re.sub(r'[^a-z0-9_]', '', safe_category)
        new_column_name = f"{column_name}_{safe_category}"
        
        if new_column_name in df.columns:
            # Overall stats
            category_count = df[new_column_name].sum()
            category_percentage = (category_count / len(df)) * 100 if len(df) > 0 else 0
            
            # Breakdown stats
            breakdown_stats = {}
            for value in breakdown_values:
                value_str = str(value)
                value_df = df[df[breakdown_column] == value]
                if len(value_df) > 0:
                    value_count = value_df[new_column_name].sum()
                    value_percentage = (value_count / breakdown_totals[value_str]) * 100 if breakdown_totals[value_str] > 0 else 0
                    breakdown_stats[value_str] = {
                        "count": int(value_count),
                        "percentage": round(value_percentage, 1),
                        "total_in_group": breakdown_totals[value_str]
                    }
                else:
                    breakdown_stats[value_str] = {
                        "count": 0,
                        "percentage": 0.0,
                        "total_in_group": 0
                    }
            
            summary["categories"][category] = {
                "count": int(category_count),
                "percentage": round(category_percentage, 1),
                "breakdown": breakdown_stats
            }
    
    return summary
